import yaml
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from enum import IntEnum

import excel_styles_and_validation

""" Tool for exporting the MASVS requirements as a checklist including MASTG coverage.

    By Carlos Holguera

    Copyright (c) 2022 OWASP Foundation

    Permission is hereby granted, free of charge, to any person obtaining a copy
    of this software and associated documentation files (the "Software"), to deal
    in the Software without restriction, including without limitation the rights
    to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
    copies of the Software, and to permit persons to whom the Software is
    furnished to do so, subject to the following conditions:

    The above copyright notice and this permission notice shall be included in all
    copies or substantial portions of the Software.

    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
    AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
    LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
    OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
    SOFTWARE.

"""

# TODO parametrize & create a function
# TODO read sheet, col ids and cell styles (centered, left, colored, character if true, etc) from yaml

MASVS_TITLES = {
    "V1": "Architecture, Design and Threat Modeling Requirements",
    "V2": "Data Storage and Privacy Requirements",
    "V3": "Cryptography Requirements",
    "V4": "Authentication and Session Management Requirements",
    "V5": "Network Communication Requirements",
    "V6": "Platform Interaction Requirements",
    "V7": "Code Quality and Build Setting Requirements",
    "V8": "Resilience Requirements",
}

MASVS = None
LANG = ""
MASTGVERSION = ""
MASTGCOMMIT = ""
MASVSVERSION = ""
MASVSCOMMIT = ""
TEST_CASE_ALIAS = "Test Case"
STATUS_ALIAS = "Status"

class Position(IntEnum):
    ID = 2
    MSTG_ID = 3
    TEXT = 4
    L1 = 5
    L2 = 6
    R = 7
    LINK_COMMON = 8
    LINK_ANDROID = 9
    LINK_IOS = 10
    STATUS_ANDROID = 11
    STATUS_IOS = 12

WS_BASE_CONFIG = {
    "start_row": 6,
    "columns": [
        {"col": "B", "position": Position.ID, "name": "ID", "width": 10, "style": "gray_header"},
        {"col": "C", "position": Position.MSTG_ID, "name": "MASVS-ID", "width": 25, "style": "gray_header"},
        {"col": "D", "position": Position.TEXT, "name": "Detailed Verification Requirement",  "width": 80, "style": "gray_header"},
        {"col": "E", "position": Position.L1, "name": "L1", "width": 5, "style": "gray_header"},
        {"col": "F", "position": Position.L2, "name": "L2", "width": 5, "style": "gray_header"},
        {"col": "G", "position": Position.R, "name": "R", "width": 5, "style": "gray_header"},
        {"col": "H", "position": Position.LINK_COMMON, "name": "Common", "width": 10, "style": "gray_header"},
        {"col": "I", "position": Position.LINK_ANDROID, "name": "Android", "width": 10, "style": "gray_header"},
        {"col": "J", "position": Position.LINK_IOS, "name": "iOS", "width": 10, "style": "gray_header"},
        {"col": "K", "position": Position.STATUS_ANDROID, "name": "Android", "width": 10, "style": "gray_header"},
        {"col": "L", "position": Position.STATUS_IOS, "name": "iOS", "width": 10, "style": "gray_header"},
    ]
}


def write_header(ws):

    ws.row_dimensions[2].height = 65
    ws.merge_cells(start_row=2, end_row=4, start_column=2, end_column=3)

    img = Image("../../Document/Images/logo_circle.png")
    img.height = 140
    img.width = 140
    ws.add_image(img, "C2")

    img = Image("owasp-masvs/Document/images/OWASP_logo.png")
    img.height = img.height * 0.1
    img.width = img.width * 0.1
    ws.add_image(img, "H2")

    ws["D2"].value = "Mobile Application Security Verification Standard"
    ws["D2"].style = "big_title"

    ws["D3"].value = f'=HYPERLINK("https://github.com/OWASP/owasp-mastg/releases/tag/{MASTGVERSION}", "OWASP MASTG {MASTGVERSION} (commit: {MASTGCOMMIT})")'
    ws["D3"].font = Font(name=excel_styles_and_validation.FONT, color="00C0C0C0")
    ws["D4"].value = f'=HYPERLINK("https://github.com/OWASP/owasp-masvs/releases/tag/{MASVSVERSION}", "OWASP MASVS {MASVSVERSION} (commit: {MASVSCOMMIT})")'
    ws["D4"].font = Font(name=excel_styles_and_validation.FONT, color="00C0C0C0")

def set_columns_width(ws):
    for col in WS_BASE_CONFIG.get("columns"):
        ws.column_dimensions[col.get("col")].width = col.get("width")


def set_table_headers(row, ws):
    ws.merge_cells(start_row=row, start_column=Position.LINK_COMMON, end_row=row, end_column=Position.LINK_IOS)
    ws.cell(row=row, column=Position.LINK_COMMON).value = TEST_CASE_ALIAS
    ws.cell(row=row, column=Position.LINK_COMMON).style = "gray_header"

    ws.merge_cells(start_row=row, start_column=Position.STATUS_ANDROID, end_row=row, end_column=Position.STATUS_IOS)
    ws.cell(row=row, column=Position.STATUS_ANDROID).value = STATUS_ALIAS
    ws.cell(row=row, column=Position.STATUS_ANDROID).style = "gray_header"

    row = row + 1
    for col in WS_BASE_CONFIG["columns"]:
        ws.cell(row=row, column=col.get("position")).value = col.get("name")
        ws.cell(row=row, column=col.get("position")).style = col.get("style")


def write_title(ws, row, start_column, end_column, title):
    cell = ws.cell(row=row, column=start_column)
    cell.value = title
    cell.style = "underline"
    cell.alignment = excel_styles_and_validation.align_left

    ws.merge_cells(start_row=row, end_row=row, start_column=start_column, end_column=end_column)

    ws.row_dimensions[row].height = 25  # points

def write_testcase(ws, row, column, url_string):
    ws.cell(row=row, column=column).value = f'=HYPERLINK("{url_string}", "{TEST_CASE_ALIAS}")'
    ws.cell(row=row, column=column).style = "Hyperlink"
    ws.cell(row=row, column=column).alignment = excel_styles_and_validation.align_center

def get_link_for(links, type):
    for link in links:
        if type in link:
            return link
    return None

def create_security_requirements_sheet(wb):
    ws = wb.active
    ws.title = "Security Requirements"
    ws.sheet_view.showGridLines = False
    write_header(ws)
    set_columns_width(ws)

    status_cells = 'K11:L400'
    ws.conditional_formatting.add(status_cells, excel_styles_and_validation.rule_fail)
    ws.conditional_formatting.add(status_cells, excel_styles_and_validation.rule_pass)
    ws.conditional_formatting.add(status_cells, excel_styles_and_validation.rule_na)

    row = WS_BASE_CONFIG["start_row"]

    for mstg_id, req in MASVS.items():
        req_id = req["id"].split(".")
        category = req_id[0]
        subindex = req_id[1]

        if subindex == "1":
            row = row + 1

            category_id = f"V{category}"
            category_title = MASVS_TITLES[category_id]

            write_title(ws, row, Position.ID, Position.STATUS_IOS, category_title)

            row = row + 1

            set_table_headers(row, ws)

            row = row + 1

            ws.add_data_validation(excel_styles_and_validation.status_validation)

            row = row + 2

        # End header

        ws.cell(row=row, column=Position.ID).value = req["id"]
        ws.cell(row=row, column=Position.ID).style = "center"

        ws.cell(row=row, column=Position.MSTG_ID).value = mstg_id
        ws.cell(row=row, column=Position.MSTG_ID).style = "center"

        ws.cell(row=row, column=Position.TEXT).value = req["text"]
        ws.cell(row=row, column=Position.TEXT).style = "text"

        if req["L1"]:
            ws.cell(row=row, column=Position.L1).style = "blue"
        if req["L2"]:
            ws.cell(row=row, column=Position.L2).style = "green"
        if req["R"]:
            ws.cell(row=row, column=Position.R).style = "orange"

        # ws.cell(row=row, column=col_link_common).value = "N/A"
        # ws.cell(row=row, column=col_link_common).style = "gray_header"

        # ws.cell(row=row, column=col_link_android).value = "N/A"
        # ws.cell(row=row, column=col_link_android).style = "gray_header"

        # ws.cell(row=row, column=col_link_ios).value = "N/A"
        # ws.cell(row=row, column=col_link_ios).style = "gray_header"

        if req.get("links"):
            link_common = get_link_for(req["links"], "0x04")
            link_android = get_link_for(req["links"], "0x05")
            link_ios = get_link_for(req["links"], "0x06")

            if link_common:
                write_testcase(ws, row, Position.LINK_COMMON, link_common)
            if link_android:
                write_testcase(ws, row, Position.LINK_ANDROID, link_android)
            if link_ios:
                write_testcase(ws, row, Position.LINK_IOS, link_ios)

        ws.row_dimensions[row].height = 55  # points

        status_android_cell = ws.cell(row=row, column=Position.STATUS_ANDROID).coordinate
        excel_styles_and_validation.status_validation.add(status_android_cell)
        status_ios_cell = ws.cell(row=row, column=Position.STATUS_IOS).coordinate
        excel_styles_and_validation.status_validation.add(status_ios_cell)

        row = row + 1

def create_about_sheet(wb):
    ws = wb.create_sheet("About")
    ws.sheet_view.showGridLines = False
    write_header(ws)
    set_columns_width(ws)

    row = WS_BASE_CONFIG["start_row"]
    first_col = WS_BASE_CONFIG["columns"][0].get("position")
    last_col = WS_BASE_CONFIG["columns"][-1].get("position")

    row = row + 2

    write_title(ws, row, first_col, last_col, "About the Project")

    row = row + 2

    ws.cell(row=row, column=first_col).value = "The OWASP Mobile Application Security (MAS) flagship project led by Carlos Holguera and Sven Schleier defines the industry standard for mobile application security."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://owasp.org/mas/"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    row = row + 2

    ws.cell(row=row, column=first_col).value = "The OWASP MASVS (Mobile Application Security Verification Standard) is a standard that establishes the security requirements for mobile app security."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://mas.owasp.org/MASTG/"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    row = row + 2

    ws.cell(row=row, column=first_col).value = "The OWASP MASTG (Mobile Application Security Testing Guide) is a comprehensive manual for mobile app security testing and reverse engineering. It describes technical processes for verifying the controls listed in the MASVS."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://mas.owasp.org/MASVS/"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    row = row + 2

    write_title(ws, row, first_col, last_col, "Feedback")

    row = row + 2

    ws.cell(row=row, column=first_col).value = "If you have any comments or suggestions, please post them on our GitHub Discussions."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://github.com/OWASP/owasp-mastg/discussions/categories/ideas"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    row = row + 2

    write_title(ws, row, first_col, last_col, "Licence")

    row = row + 2

    ws.cell(row=row, column=first_col).value = "Copyright © 2022 The OWASP Foundation. This work is licensed under a Creative Commons Attribution-ShareAlike 4.0 International License. For any reuse or distribution, you must make clear to others the license terms of this work."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://github.com/OWASP/owasp-mastg/blob/master/License.md"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'


def generate_spreadsheet(output_file):

    wb = Workbook()
    excel_styles_and_validation.load_styles(wb)

    create_security_requirements_sheet(wb)
    create_about_sheet(wb)

    wb.save(filename=output_file)


def main():
    global MASVS, LANG, MASTGVERSION, MASTGCOMMIT, MASVSVERSION, MASVSCOMMIT
    import argparse

    parser = argparse.ArgumentParser(description="Export the MASVS requirements as Excel. Default language is en.")
    parser.add_argument("-m", "--masvs", required=True)
    parser.add_argument("-l", "--lang", required=True)
    parser.add_argument("-o", "--outputfile", required=True)
    parser.add_argument("-v1", "--mastgversion", required=True)
    parser.add_argument("-c1", "--mastgcommit", required=True)
    parser.add_argument("-v2", "--masvsversion", required=True)
    parser.add_argument("-c2", "--masvscommit", required=True)

    args = parser.parse_args()

    # set global vars
    MASVS = yaml.safe_load(open(args.masvs))
    LANG = args.lang
    MASTGVERSION = args.mastgversion
    MASTGCOMMIT = args.mastgcommit
    MASVSVERSION = args.masvsversion
    MASVSCOMMIT = args.masvscommit

    print(f"Generating {LANG.upper()} Checklist for MASTG {MASTGVERSION} ({MASTGCOMMIT}) and MASVS {MASVSVERSION} ({MASVSCOMMIT})")

    generate_spreadsheet(args.outputfile)


if __name__ == "__main__":
    main()
